#!/usr/bin/env python3
"""Generate the KClinics Financial Model workbook from workflow-extracted inputs.

Tabs: README, Assumptions, Machinery, Energy, Pricing, Volumes, P&L, Data.
Conventions: blue font = editable input, yellow fill = key assumption to review,
black = formula, green = cross-sheet link. Arial throughout.
"""
import json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

import os
HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "model-inputs.json")
OUT = os.path.join(HERE, "KClinics-Financial-Model.xlsx")

data = json.load(open(SRC))
syn = data["synthesis"]
cat = data["catalogue"]
rows_t = [t for t in syn["treatments"] if not t.get("meta")]
mach = [m for m in syn["machinery"] if not m.get("meta") and m["machineClass"] != "dental-surgery-DEFERRED"]
energy = syn["energy"]
opex = syn["opex"]
tax = syn["tax"]
gaps = syn["gaps"]

MONTHS = 36
EXTRACT_DATE = "2026-08-28"

# ---------- styles ----------
F_TITLE = Font(name="Arial", size=16, bold=True, color="1F3B4D")
F_H1 = Font(name="Arial", size=12, bold=True, color="FFFFFF")
F_H2 = Font(name="Arial", size=10, bold=True)
F_LBL = Font(name="Arial", size=10)
F_IN = Font(name="Arial", size=10, color="0000FF")          # blue input
F_FM = Font(name="Arial", size=10)                           # black formula
F_LINK = Font(name="Arial", size=10, color="008000")         # green cross-sheet
F_NOTE = Font(name="Arial", size=8, italic=True, color="666666")
F_WARN = Font(name="Arial", size=9, bold=True, color="B00000")
FILL_H1 = PatternFill("solid", fgColor="1F3B4D")
FILL_SEC = PatternFill("solid", fgColor="DCE6F1")
FILL_IN = PatternFill("solid", fgColor="FFF8DC")             # pale key-assumption fill
FILL_TOT = PatternFill("solid", fgColor="EAEFF5")
FILL_KPI = PatternFill("solid", fgColor="E2EFDA")
THIN = Border(bottom=Side(style="thin", color="B0B0B0"))
GBP2 = '"£"#,##0.00'
GBP0 = '"£"#,##0;("£"#,##0);"-"'
PCT1 = "0.0%"
PCT0 = "0%"
NUM1 = "#,##0.0"
NUM0 = "#,##0"

def sec(ws, row, text, width=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_H2
    c.fill = FILL_SEC
    for k in range(2, width + 1):
        ws.cell(row=row, column=k).fill = FILL_SEC
    return row + 1

wb = Workbook()

# =====================================================================
# Assumptions
# =====================================================================
ws = wb.active
ws.title = "Assumptions"
A = {}  # registry: key -> absolute ref on Assumptions

ws["A1"] = "Assumptions — every editable input in the model"
ws["A1"].font = F_TITLE
ws["A2"] = "Blue on cream = edit me. Column D says where each number came from. Nothing else in the workbook is hardcoded."
ws["A2"].font = F_NOTE
for col, w in zip("ABCD", (44, 14, 12, 95)):
    ws.column_dimensions[col].width = w

r = 4
def put(ws, r, label, value, unit, basis, key=None, fmt=None, formula=False):
    ws.cell(row=r, column=1, value=label).font = F_LBL
    c = ws.cell(row=r, column=2, value=value)
    c.font = F_FM if formula else F_IN
    if not formula:
        c.fill = FILL_IN
    if fmt:
        c.number_format = fmt
    ws.cell(row=r, column=3, value=unit).font = F_NOTE
    b = ws.cell(row=r, column=4, value=basis)
    b.font = F_NOTE
    b.alignment = Alignment(wrap_text=True, vertical="top")
    if key:
        A[key] = f"Assumptions!$B${r}"
    return r + 1

r = sec(ws, r, "CLINIC & CAPACITY", 4)
r = put(ws, r, "Model start month", datetime.date(2026, 9, 1), "date", "First forecast month. Clinic re-opened 30 May 2026; model starts Sep 2026.", "start", "mmm yyyy")
r = put(ws, r, "Usable treatment rooms", 9, "rooms", "11 active rooms in DB (3 Aesthetics, 2 Dental, 6 Treatment); 2 dental rooms excluded until dentistry launches (separate company). ESTIMATE — confirm dental rooms are not used as overflow.", "rooms", NUM0)
r = put(ws, r, "Rooms heated/cooled (HVAC)", 9, "rooms", "Rooms actively conditioned. Set to 11 if dental rooms are kept conditioned while mothballed.", "hvac_rooms", NUM0)
r = put(ws, r, "Open days per week", 6, "days", "Live booking hours imply ~57 open hours/week.", "open_days", NUM0)
r = put(ws, r, "Open hours per day", 9.5, "hours", "57 open hours/week ÷ 6 days.", "open_hours", NUM1)
r = put(ws, r, "Assumed room size", 15, "m²", "ASSUMPTION — no floor area in codebase or lease data. Drives the HVAC curve on the Energy tab.", "room_m2", NUM0)

r = sec(ws, r, "STAFFING", 4)
r = put(ws, r, "Practitioner salary", opex["staff"]["practitionerSalaryGBP"], "£/yr", opex["staff"]["practitionerBasis"], "sal_prac", GBP0)
r = put(ws, r, "Receptionist salary", opex["staff"]["receptionistSalaryGBP"], "£/yr", opex["staff"]["receptionistBasis"], "sal_rec", GBP0)
r = put(ws, r, "Owner salary (market rate)", 48000, "£/yr", "GAP — owner remuneration undefined. £48k market-rate placeholder so owner labour is not modelled as free. Set to your actual salary/dividend policy.", "sal_owner", GBP0)
r = put(ws, r, "Employer on-cost multiplier", 1.16, "×", opex["staff"]["onCostBasis"], "oncost", "0.00")
r = put(ws, r, "Practitioners — headcount Y1 / Y2 / Y3", 2, "FTE", "Y1: 1 employed practitioner + working owner treating. Scale with the utilisation ramp — the capacity check at the bottom of P&L warns when treated hours exceed practitioner capacity.", "hc_prac_y1", NUM0)
ws.cell(row=r-1, column=3, value=3).font = F_IN; ws.cell(row=r-1, column=3).fill = FILL_IN; ws.cell(row=r-1, column=3).number_format = NUM0
ws.cell(row=r-1, column=5, value=5).font = F_IN; ws.cell(row=r-1, column=5).fill = FILL_IN; ws.cell(row=r-1, column=5).number_format = NUM0
A["hc_prac_y2"] = f"Assumptions!$C${r-1}"; A["hc_prac_y3"] = f"Assumptions!$E${r-1}"
# fix: unit text got overwritten by Y2 headcount; restore basis col D
ws.cell(row=r-1, column=4, value="Y1 in column B, Y2 in column C, Y3 in column E. Defaults 2/3/5 track the default utilisation ramp.").font = F_NOTE
r = put(ws, r, "Receptionists — headcount Y1 / Y2 / Y3", 1, "FTE", "Y1 in B, Y2 in C, Y3 in E.", "hc_rec_y1", NUM0)
ws.cell(row=r-1, column=3, value=1).font = F_IN; ws.cell(row=r-1, column=3).fill = FILL_IN; ws.cell(row=r-1, column=3).number_format = NUM0
ws.cell(row=r-1, column=5, value=2).font = F_IN; ws.cell(row=r-1, column=5).fill = FILL_IN; ws.cell(row=r-1, column=5).number_format = NUM0
A["hc_rec_y2"] = f"Assumptions!$C${r-1}"; A["hc_rec_y3"] = f"Assumptions!$E${r-1}"
ws.cell(row=r-1, column=4, value="Y1 in column B, Y2 in column C, Y3 in column E.").font = F_NOTE
r = put(ws, r, "Practitioner productive treatment hours/week", 30, "hrs", "Hours actually treating (vs admin/gaps) — drives labour cost per session in Pricing.", "prod_hrs", NUM0)
r = put(ws, r, "Working weeks per year", 46, "wks", "52 minus holiday/sickness/training.", "work_wks", NUM0)
r = put(ws, r, "Loaded practitioner cost per treatment hour",
        f"={A['sal_prac']}*{A['oncost']}/({A['work_wks']}*{A['prod_hrs']})", "£/hr",
        "Calculated: salary × on-cost ÷ (working weeks × productive hours). Used in Pricing unit economics.", "labour_hr", GBP2, formula=True)
r = put(ws, r, "Practitioner capacity per month",
        f"={A['prod_hrs']}*{A['work_wks']}/12", "hrs/mo", "Per practitioner FTE — used by the P&L capacity check.", "prac_cap", NUM1, formula=True)

r = sec(ws, r, "ENERGY", 4)
r = put(ws, r, "Electricity unit rate", energy["unitRatePencePerKWh"], "p/kWh", energy["unitRateBasis"], "kwh_rate", NUM1)
r = put(ws, r, "Standing charge", energy["standingChargePencePerDay"], "p/day", energy["standingChargeBasis"], "standing", NUM1)
r = put(ws, r, "Base load (lighting, fridge, autoclave, PCs)", energy["baseLoadKWhPerDay"], "kWh/day", energy["baseLoadBasis"], "baseload", NUM1)
r = put(ws, r, "Heating & cooling", "see Energy tab", "", energy["gasAssumption"] + " Seasonal per-room kWh curve is editable on the Energy tab.", None, formula=True)

r = sec(ws, r, "PROPERTY", 4)
r = put(ws, r, "Rent", opex["rent"]["placeholderGBPPerYear"], "£/yr", opex["rent"]["guidance"], "rent", GBP0)
r = put(ws, r, "Business rates", opex["businessRates"]["placeholderGBPPerYear"], "£/yr", opex["businessRates"]["guidance"], "rates", GBP0)
r = put(ws, r, "Service charge / building costs", 0, "£/yr", "Unknown — set from the lease.", "svc_charge", GBP0)

r = sec(ws, r, "OTHER OVERHEADS", 4)
r = put(ws, r, "Insurance (treatment risk + PL + laser cover)", opex["insuranceGBPPerYear"], "£/yr", opex["insuranceBasis"], "insurance", GBP0)
r = put(ws, r, "Platform hosting & infra", opex["softwareGBPPerMonth"], "£/mo", opex["softwareBasis"], "software", GBP0)
r = put(ws, r, "Card processing fee", opex["cardFeesPctOfCardTurnover"] / 100, "% of card takings", opex["cardFeesBasis"], "card_fee", PCT1)
r = put(ws, r, "Card share of turnover", 0.95, "%", "Near-cashless assumption (Stripe card-on-file flow in use).", "card_share", PCT0)
r = put(ws, r, "Cleaning & sundries", 400, "£/mo", "ESTIMATE — laundry, room consumables not tied to treatments, waste disposal (clinical waste contract).", "cleaning", GBP0)
r = put(ws, r, "Other fixed costs", 200, "£/mo", "ESTIMATE — accountancy, bank, phone/VoIP, misc subscriptions.", "other_fixed", GBP0)

r = sec(ws, r, "MARKETING", 4)
r = put(ws, r, "Marketing % of net revenue — Y1 / Y2 / Y3", 0.12, "%", opex["marketing"]["basis"] + " Y1 in B, Y2 in C, Y3 in E.", "mkt_y1", PCT0)
ws.cell(row=r-1, column=3, value=0.10).font = F_IN; ws.cell(row=r-1, column=3).fill = FILL_IN; ws.cell(row=r-1, column=3).number_format = PCT0
ws.cell(row=r-1, column=5, value=0.08).font = F_IN; ws.cell(row=r-1, column=5).fill = FILL_IN; ws.cell(row=r-1, column=5).number_format = PCT0
A["mkt_y2"] = f"Assumptions!$C${r-1}"; A["mkt_y3"] = f"Assumptions!$E${r-1}"
r = put(ws, r, "Marketing monthly floor", opex["marketing"]["fixedFloorGBPPerMonth"], "£/mo", "Spend floor while revenue is small — you cannot ramp without visibility. CAC benchmark ~£45/booked consultation.", "mkt_floor", GBP0)

r = sec(ws, r, "VAT & TAX", 4)
r = put(ws, r, "VAT standard rate", 0.20, "%", tax["vat"]["aesthetics"], "vat_rate", PCT0)
r = put(ws, r, "VAT registration threshold", 90000, "£ rolling 12-mo", tax["vat"]["registrationStatus"], "vat_thresh", GBP0)
r = put(ws, r, "Taxable turnover in the 12 months before model start", 25000, "£", "ESTIMATE from 225 bookings in the 90 days post-reopening — feeds the rolling-12-month VAT threshold tracker in P&L. Replace with the real figure from your accounts.", "vat_prior", GBP0)
r = put(ws, r, "VAT registration month override", 0, "month # (0 = auto)", "0 = the model registers you automatically the month after rolling 12-mo turnover crosses the threshold. Set 1–36 to force a date; prices are held (VAT absorbed out of margin) — see README.", "vat_override", NUM0)
r = put(ws, r, "Corporation tax — small profits rate / limit", 0.19, "%", f"2026/27: 19% to £{tax['corporationTax']['smallProfitsUpToGBP']:,}; 26.5% effective £50k–£250k (marginal relief); 25% main rate. Annual columns in P&L apply the bands properly.", "ct_small", PCT0)
ws.cell(row=r-1, column=3, value=50000).font = F_IN; ws.cell(row=r-1, column=3).fill = FILL_IN; ws.cell(row=r-1, column=3).number_format = GBP0
A["ct_small_lim"] = f"Assumptions!$C${r-1}"
r = put(ws, r, "Corporation tax — marginal band rate / upper limit", 0.265, "%", "Effective rate between the limits via marginal relief.", "ct_marg", PCT1)
ws.cell(row=r-1, column=3, value=250000).font = F_IN; ws.cell(row=r-1, column=3).fill = FILL_IN; ws.cell(row=r-1, column=3).number_format = GBP0
A["ct_marg_lim"] = f"Assumptions!$C${r-1}"
r = put(ws, r, "Corporation tax — main rate", 0.25, "%", tax["aia"]["note"], "ct_main", PCT0)

# =====================================================================
# Machinery
# =====================================================================
ws = wb.create_sheet("Machinery")
M = {}
ws["A1"] = "Machinery, enhancements, depreciation & amortisation"
ws["A1"].font = F_TITLE
ws["A2"] = ("Purchase prices are class-typical UK market figures (see basis column) — replace with your invoice amounts. "
            "In-service month: 0 = owned before the model starts; 1–36 = bought in that forecast month; blank = not owned (excluded). "
            "Planned hours/month is only the allocation base for service & depreciation per treatment hour in Pricing.")
ws["A2"].font = F_NOTE
headers = ["Machine", "Class key", "Ownership", "Purchase £", "In-service month", "Life (yrs)", "Monthly depreciation",
           "kWh per treatment hr", "Annual service £", "Planned hours/mo", "Service £/hr", "Depreciation £/hr", "Consumables model", "Basis"]
HR = 4
for i, h in enumerate(headers, 1):
    c = ws.cell(row=HR, column=i, value=h)
    c.font = F_H1
    c.fill = FILL_H1
    c.alignment = Alignment(wrap_text=True, vertical="center")
for col, w in zip("ABCDEFGHIJKLMN", (34, 20, 26, 12, 12, 9, 13, 12, 12, 12, 11, 13, 44, 44)):
    ws.column_dimensions[col].width = w

planned_hours = {
    "diode-laser": 250, "laser-ipl-platform": 30, "pico-laser-PLANNED": 25, "co2-laser": 8,
    "hifu": 25, "rf-tightening": 12, "hydrafacial-platform": 60, "endosphere-vacuum": 45,
    "microneedling-pen": 18, "microdermabrasion": 8, "microcurrent-led": 25,
    "none-injectables": 25, "none-manual": 25,
}
mrow = HR + 1
first_m, last_m = mrow, mrow + len(mach) - 1
for m in mach:
    key = m["machineClass"]
    planned = m["ownership"].upper().find("NOT") == 0 or "PLANNED" in key
    ws.cell(row=mrow, column=1, value=m["name"]).font = F_LBL
    ws.cell(row=mrow, column=2, value=key).font = F_LBL
    ow = ws.cell(row=mrow, column=3, value=m["ownership"])
    ow.font = F_WARN if ("VERIFY" in m["ownership"] or "AMBIG" in m["ownership"].upper() or planned) else F_NOTE
    ow.alignment = Alignment(wrap_text=True, vertical="top")
    for col, val, fmt in ((4, m["purchaseGBP"], GBP0), (6, m["lifeYears"], NUM0),
                          (8, m["kWhPerTreatmentHour"], "0.00"), (9, m["annualServiceGBP"], GBP0),
                          (10, planned_hours.get(key, 20), NUM0)):
        c = ws.cell(row=mrow, column=col, value=val)
        c.font = F_IN
        c.fill = FILL_IN
        c.number_format = fmt
    insv = ws.cell(row=mrow, column=5, value=None if planned else 0)
    insv.font = F_IN
    insv.fill = FILL_IN
    insv.number_format = NUM0
    dep = ws.cell(row=mrow, column=7, value=f"=IF(E{mrow}=\"\",0,D{mrow}/(F{mrow}*12))")
    dep.font = F_FM
    dep.number_format = GBP2
    ws.cell(row=mrow, column=11, value=f"=IF(J{mrow}=0,0,I{mrow}/12/J{mrow})").number_format = GBP2
    ws.cell(row=mrow, column=12, value=f"=IF(J{mrow}=0,0,G{mrow}/J{mrow})").number_format = GBP2
    cns = ws.cell(row=mrow, column=13, value=m["consumablesModel"])
    cns.font = F_NOTE
    cns.alignment = Alignment(wrap_text=True, vertical="top")
    bas = ws.cell(row=mrow, column=14, value=m["basis"])
    bas.font = F_NOTE
    bas.alignment = Alignment(wrap_text=True, vertical="top")
    M[key] = mrow
    mrow += 1

# Enhancements block
er = mrow + 2
ws.cell(row=er, column=1, value="ENHANCEMENTS & FIT-OUT (amortised)").font = F_H2
ws.cell(row=er, column=1).fill = FILL_SEC
for k in range(2, 8):
    ws.cell(row=er, column=k).fill = FILL_SEC
er += 1
for i, h in enumerate(["Item", "", "Notes", "Cost £", "In-service month", "Life (yrs)", "Monthly amortisation"], 1):
    if h:
        ws.cell(row=er, column=i, value=h).font = F_H2
er += 1
enh_first = er
enh_rows = [
    ("Initial fit-out / leasehold improvements", "GAP — fit-out spend on 4 Charterhouse Buildings unknown. Enter the capitalised amount; it amortises over the life you set.", 0, 0, 10),
    ("EXAMPLE — reception refresh (edit or zero out)", "Example row showing the expected format.", 8000, 6, 5),
    ("Planned enhancement 2", "", 0, None, 5),
    ("Planned enhancement 3", "", 0, None, 5),
    ("Planned enhancement 4", "", 0, None, 5),
]
for name, note, cost, insv, life in enh_rows:
    ws.cell(row=er, column=1, value=name).font = F_LBL
    n = ws.cell(row=er, column=3, value=note)
    n.font = F_NOTE
    n.alignment = Alignment(wrap_text=True, vertical="top")
    for col, val, fmt in ((4, cost, GBP0), (5, insv, NUM0), (6, life, NUM0)):
        c = ws.cell(row=er, column=col, value=val)
        c.font = F_IN
        c.fill = FILL_IN
        c.number_format = fmt
    a = ws.cell(row=er, column=7, value=f"=IF(OR(E{er}=\"\",F{er}=0),0,D{er}/(F{er}*12))")
    a.font = F_FM
    a.number_format = GBP2
    er += 1
enh_last = er - 1

# Depreciation & amortisation schedule (36 months)
sr = er + 2
ws.cell(row=sr, column=1, value="MONTHLY D&A SCHEDULE (feeds P&L)").font = F_H2
ws.cell(row=sr, column=1).fill = FILL_SEC
sr += 1
ws.cell(row=sr, column=1, value="Month #").font = F_H2
for mth in range(1, MONTHS + 1):
    c = ws.cell(row=sr, column=1 + mth, value=mth)
    c.font = F_H2
    c.number_format = NUM0
sched_hdr = sr
sr += 1
dep_total_row = None
for label, rows, valcol in (("mach", range(first_m, last_m + 1), "G"), ("enh", range(enh_first, enh_last + 1), "G7ENH")):
    for rr in rows:
        ws.cell(row=sr, column=1, value=f"='{ws.title}'!A{rr}" if False else None)
        ws.cell(row=sr, column=1, value=f"=A{rr}").font = F_NOTE
        for mth in range(1, MONTHS + 1):
            col = gcl(1 + mth)
            if label == "mach":
                f = (f"=IF($E{rr}=\"\",0,IF(AND(COLUMN()-1>=MAX($E{rr},1),COLUMN()-1<$E{rr}+$F{rr}*12),$G{rr},0))")
            else:
                f = (f"=IF(OR($E{rr}=\"\",$F{rr}=0),0,IF(AND(COLUMN()-1>=MAX($E{rr},1),COLUMN()-1<$E{rr}+$F{rr}*12),$G{rr},0))")
            c = ws.cell(row=sr, column=1 + mth, value=f)
            c.font = F_FM
            c.number_format = GBP0
        if label == "mach":
            M.setdefault("_sched", {})[rr] = sr
        sr += 1
# totals
dep_row = sr
ws.cell(row=sr, column=1, value="Total depreciation (machines)").font = F_H2
n_mach = last_m - first_m + 1
for mth in range(1, MONTHS + 1):
    col = gcl(1 + mth)
    c = ws.cell(row=sr, column=1 + mth,
                value=f"=SUM({col}{sched_hdr+1}:{col}{sched_hdr+n_mach})")
    c.font = F_FM
    c.fill = FILL_TOT
    c.number_format = GBP0
sr += 1
amort_row = sr
ws.cell(row=sr, column=1, value="Total amortisation (enhancements)").font = F_H2
n_enh = enh_last - enh_first + 1
for mth in range(1, MONTHS + 1):
    col = gcl(1 + mth)
    c = ws.cell(row=sr, column=1 + mth,
                value=f"=SUM({col}{sched_hdr+1+n_mach}:{col}{sched_hdr+n_mach+n_enh})")
    c.font = F_FM
    c.fill = FILL_TOT
    c.number_format = GBP0

MACH = {"dep_row": dep_row, "amort_row": amort_row, "rows": M, "first": first_m, "last": last_m}

# =====================================================================
# Pricing
# =====================================================================
ws = wb.create_sheet("Pricing")
ws["A1"] = "Treatment pricing architecture — unit economics per variant"
ws["A1"].font = F_TITLE
ws["A2"] = ("Every recommended price is a formula: fully-loaded cost ÷ (1 − target margin − card-fee drag), then VAT, then rounding. "
            "Change the levers below or any machine/labour assumption and the whole price list moves. "
            "Current prices scraped from the live admin catalogue on " + EXTRACT_DATE + ". Recorded COGS is empty on every variant, so consumables are researched estimates (blue).")
ws["A2"].font = F_NOTE

# Lever block
LV = {}
lr = 4
ws.cell(row=lr, column=1, value="LEVERS").font = F_H2
ws.cell(row=lr, column=1).fill = FILL_SEC
for k in range(2, 7):
    ws.cell(row=lr, column=k).fill = FILL_SEC
lr += 1
def lever(lrow, label, val, fmt, note, key):
    ws.cell(row=lrow, column=1, value=label).font = F_LBL
    c = ws.cell(row=lrow, column=2, value=val)
    c.font = F_IN
    c.fill = FILL_IN
    c.number_format = fmt
    n = ws.cell(row=lrow, column=3, value=note)
    n.font = F_NOTE
    n.alignment = Alignment(wrap_text=True)
    LV[key] = f"$B${lrow}"
    return lrow + 1
lr = lever(lr, "Price for the VAT-registered state? (1 = yes)", 1, NUM0,
           "1 = recommended prices keep their margin AFTER VAT registration (the safe basis for a durable price list). 0 = price for today's unregistered state.", "vat_on")
lr = lever(lr, "Round recommended prices to nearest £", 1, GBP0, "1, 5 or 9 — e.g. 5 gives £85/£90-style pricing.", "round_to")
lr = lever(lr, "Course discount — 3 sessions", 0.10, PCT0, "Per-session discount vs single price. Current live courses average ~15% off.", "disc3")
lr = lever(lr, "Course discount — 6 sessions", 0.15, PCT0, "", "disc6")
lr = lever(lr, "Course discount — 10 sessions", 0.20, PCT0, "", "disc10")

# target margin table by class
tr = lr + 1
ws.cell(row=tr, column=1, value="TARGET CONTRIBUTION MARGIN BY MACHINE CLASS").font = F_H2
ws.cell(row=tr, column=1).fill = FILL_SEC
for k in range(2, 7):
    ws.cell(row=tr, column=k).fill = FILL_SEC
tr += 1
ws.cell(row=tr, column=1, value="Class").font = F_H2
ws.cell(row=tr, column=2, value="Target margin").font = F_H2
ws.cell(row=tr, column=3, value="Margin = (net revenue − fully-loaded cost) ÷ net revenue. Fully-loaded cost includes consumables, machine energy, machine service & depreciation per hour, practitioner time and card fees.").font = F_NOTE
tr += 1
target_default = {
    "none-injectables": 0.55, "hifu": 0.60, "hydrafacial-platform": 0.65,
    "microneedling-pen": 0.65, "none-manual": 0.65,
}
class_target_first = tr
class_keys = [m["machineClass"] for m in mach]
CLS_ROW = {}
for key in class_keys:
    ws.cell(row=tr, column=1, value=key).font = F_LBL
    c = ws.cell(row=tr, column=2, value=target_default.get(key, 0.70))
    c.font = F_IN
    c.fill = FILL_IN
    c.number_format = PCT0
    CLS_ROW[key] = tr
    tr += 1
class_target_last = tr - 1
TGT_RANGE = f"$A${class_target_first}:$A${class_target_last}"
TGT_VALS = f"$B${class_target_first}:$B${class_target_last}"

# main table
hr = tr + 2
cols = ["Service", "Variant", "Category", "Status", "Machine class", "Internal min", "Machine min", "VAT %",
        "Current price £", "Current net £", "Consumables £", "Machine energy £", "Machine service £",
        "Machine depreciation £", "Practitioner £", "Card fee £", "Fully-loaded cost £", "Contribution £",
        "Contribution %", "Target %", "Recommended net £", "Recommended price £", "Δ vs current",
        "Course 3 rec £", "Course 6 rec £", "Course 10 rec £", "Current course 6 £", "Current 6-course discount", "Sellable"]
for i, h in enumerate(cols, 1):
    c = ws.cell(row=hr, column=i, value=h)
    c.font = F_H1
    c.fill = FILL_H1
    c.alignment = Alignment(wrap_text=True, vertical="center")
widths = [30, 34, 10, 13, 18, 8, 8, 7, 10, 10, 11, 10, 10, 12, 11, 9, 12, 11, 10, 8, 12, 13, 9, 10, 10, 10, 10, 10, 7]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[gcl(i)].width = w
ws.freeze_panes = "C1"

pr = hr + 1
p_first = pr
mach_sheet = "Machinery"
for t in rows_t:
    key = t["machineClass"]
    mr_ = MACH["rows"].get(key)
    ws.cell(row=pr, column=1, value=t["service"]).font = F_LBL
    ws.cell(row=pr, column=2, value=t["variant"]).font = F_LBL
    ws.cell(row=pr, column=3, value=t["category"]).font = F_NOTE
    st = ws.cell(row=pr, column=4, value=t["status"])
    st.font = F_WARN if t["status"] in ("INACTIVE", "COMING_SOON") else F_NOTE
    ws.cell(row=pr, column=5, value=key).font = F_NOTE
    ws.cell(row=pr, column=6, value=t["durMin"]).number_format = NUM0
    ws.cell(row=pr, column=7, value=t.get("machineMin") or 0).number_format = NUM0
    ws.cell(row=pr, column=8, value=(t.get("vatPct") or 0) / 100).number_format = PCT0
    ws.cell(row=pr, column=9, value=t["priceGBP"]).number_format = GBP2
    ws.cell(row=pr, column=10, value=f"=I{pr}/(1+H{pr}*{LV['vat_on']})").number_format = GBP2
    ccell = ws.cell(row=pr, column=11, value=t.get("consumablesGBP") or 0)
    ccell.font = F_IN
    ccell.fill = FILL_IN
    ccell.number_format = GBP2
    if mr_:
        ws.cell(row=pr, column=12, value=f"=G{pr}/60*{mach_sheet}!$H${mr_}*{A['kwh_rate']}/100").number_format = GBP2
        ws.cell(row=pr, column=13, value=f"=G{pr}/60*{mach_sheet}!$K${mr_}").number_format = GBP2
        ws.cell(row=pr, column=14, value=f"=G{pr}/60*{mach_sheet}!$L${mr_}").number_format = GBP2
    else:
        ws.cell(row=pr, column=12, value=0).number_format = GBP2
        ws.cell(row=pr, column=13, value=0).number_format = GBP2
        ws.cell(row=pr, column=14, value=0).number_format = GBP2
    ws.cell(row=pr, column=15, value=f"=F{pr}/60*{A['labour_hr']}").number_format = GBP2
    ws.cell(row=pr, column=16, value=f"=I{pr}*{A['card_fee']}*{A['card_share']}").number_format = GBP2
    ws.cell(row=pr, column=17, value=f"=SUM(K{pr}:P{pr})").number_format = GBP2
    ws.cell(row=pr, column=18, value=f"=J{pr}-Q{pr}").number_format = GBP2
    ws.cell(row=pr, column=19, value=f"=IF(J{pr}=0,0,R{pr}/J{pr})").number_format = PCT1
    ws.cell(row=pr, column=20, value=f"=INDEX({TGT_VALS},MATCH(E{pr},{TGT_RANGE},0))").number_format = PCT0
    # net = non-card cost / (1 - target - cardfee*share*(1+vat))
    ws.cell(row=pr, column=21,
            value=(f"=SUM(K{pr}:O{pr})/(1-T{pr}-{A['card_fee']}*{A['card_share']}*(1+H{pr}*{LV['vat_on']}))")
            ).number_format = GBP2
    ws.cell(row=pr, column=22,
            value=f"=MROUND(U{pr}*(1+H{pr}*{LV['vat_on']}),{LV['round_to']})").number_format = GBP2
    ws.cell(row=pr, column=23, value=f"=IF(I{pr}=0,0,V{pr}/I{pr}-1)").number_format = PCT0
    ws.cell(row=pr, column=24, value=f"=MROUND(V{pr}*3*(1-{LV['disc3']}),{LV['round_to']})").number_format = GBP0
    ws.cell(row=pr, column=25, value=f"=MROUND(V{pr}*6*(1-{LV['disc6']}),{LV['round_to']})").number_format = GBP0
    ws.cell(row=pr, column=26, value=f"=MROUND(V{pr}*10*(1-{LV['disc10']}),{LV['round_to']})").number_format = GBP0
    cur6 = None
    for co in (t.get("courses") or []):
        if co.get("sessions") == 6:
            cur6 = co["totalPence"] / 100
    if cur6 is not None:
        ws.cell(row=pr, column=27, value=cur6).number_format = GBP0
        ws.cell(row=pr, column=28, value=f"=1-AA{pr}/(6*I{pr})").number_format = PCT0
    ws.cell(row=pr, column=29, value=f"=IF(OR(D{pr}=\"ACTIVE\",D{pr}=\"CONSULTATION\"),1,0)").number_format = NUM0
    pr += 1
p_last = pr - 1
PRC = {"first": p_first, "last": p_last}

# =====================================================================
# Volumes
# =====================================================================
ws = wb.create_sheet("Volumes")
ws["A1"] = "Volume & revenue forecast — capacity × utilisation × mix"
ws["A1"].font = F_TITLE
ws["A2"] = ("Room capacity comes from Assumptions. The utilisation ramp (blue row) is THE dominant free assumption — "
            "current run-rate is ~3% (225 bookings in the 90 days after the May 2026 reopening). "
            "Mix % splits treated hours across machine classes; seeded from live booking mix. "
            "Per-class average price/cost auto-computed from the Pricing tab (sellable variants); use the blue override column to demand-weight.")
ws["A2"].font = F_NOTE

fam_defs = []  # (label, class_key, seed_mix, dur_override, price_override, cost_override)
book_mix = {  # bookings90d mapped to class
    "diode-laser": 141, "consultation": 38, "hydrafacial-platform": 14, "endosphere-vacuum": 10,
    "microcurrent-led": 6, "microneedling-pen": 4, "rf-tightening": 3, "laser-ipl-platform": 5,
    "microdermabrasion": 1, "none-injectables": 2, "hifu": 1, "none-manual": 0, "co2-laser": 0,
    "pico-laser-PLANNED": 0,
}
avg_dur = {"diode-laser": 25, "consultation": 20, "hydrafacial-platform": 50, "endosphere-vacuum": 50,
           "microcurrent-led": 45, "microneedling-pen": 45, "rf-tightening": 45, "laser-ipl-platform": 30,
           "microdermabrasion": 40, "none-injectables": 30, "hifu": 60, "none-manual": 45, "co2-laser": 30,
           "pico-laser-PLANNED": 25}
hours_w = {k: book_mix[k] * avg_dur[k] for k in book_mix}
tot_w = sum(hours_w.values())
seed_mix = {k: round(v / tot_w * 200) / 2 for k, v in hours_w.items()}  # to 0.5%
# small floors for families with tiny/no bookings (planned growth), rebalance from diode
floors = {"none-injectables": 3.0, "hifu": 2.0, "laser-ipl-platform": 3.0, "none-manual": 2.0,
          "rf-tightening": 1.0, "microneedling-pen": 2.0, "co2-laser": 0.0, "pico-laser-PLANNED": 0.0,
          "microdermabrasion": 0.5}
for k, fl in floors.items():
    if seed_mix.get(k, 0) < fl:
        seed_mix[k] = fl
diff = 100 - sum(seed_mix.values())
seed_mix["diode-laser"] = round((seed_mix["diode-laser"] + diff) * 2) / 2

fam_order = ["diode-laser", "laser-ipl-platform", "hifu", "rf-tightening", "hydrafacial-platform",
             "endosphere-vacuum", "microneedling-pen", "microdermabrasion", "microcurrent-led",
             "none-injectables", "none-manual", "co2-laser", "pico-laser-PLANNED", "consultation"]
fam_labels = {
    "diode-laser": "Laser hair removal (diode)", "laser-ipl-platform": "IPL / laser skin",
    "hifu": "HIFU lifting", "rf-tightening": "RF tightening", "hydrafacial-platform": "HydraFacial & facials",
    "endosphere-vacuum": "Body contouring (Endosphere/vacuum)", "microneedling-pen": "Microneedling & BB Glow",
    "microdermabrasion": "Microdermabrasion", "microcurrent-led": "Microcurrent & LED",
    "none-injectables": "Injectables", "none-manual": "Peels, massage & manual",
    "co2-laser": "CO2 laser (intimate) — coming soon", "pico-laser-PLANNED": "Tattoo/pigment removal (pico) — not yet purchased",
    "consultation": "Consultations (no revenue)",
}

# Family table
fhr = 4
fam_cols = ["Family", "Class key", "Avg session min (auto)", "Min override", "Min used",
            "Avg price £ (auto)", "Price override", "Price used £",
            "Avg direct cost £ (auto)", "Cost override", "Cost used £",
            "Machine min/session", "kWh/treatment hr", "Hours mix %"]
for i, h in enumerate(fam_cols, 1):
    c = ws.cell(row=fhr, column=i, value=h)
    c.font = F_H1
    c.fill = FILL_H1
    c.alignment = Alignment(wrap_text=True, vertical="center")
for i, w in enumerate([32, 20, 11, 9, 8, 11, 9, 10, 11, 9, 9, 11, 10, 9], 1):
    ws.column_dimensions[gcl(i)].width = w

fr = fhr + 1
f_first = fr
FAM_ROW = {}
PF, PL = PRC["first"], PRC["last"]
prc = "Pricing"
for key in fam_order:
    lbl = fam_labels[key]
    ws.cell(row=fr, column=1, value=lbl).font = F_LBL
    ws.cell(row=fr, column=2, value=key).font = F_NOTE
    if key == "consultation":
        ws.cell(row=fr, column=3, value=0).number_format = NUM1
        for col, v in ((4, 20), (7, 0), (10, 1)):
            c = ws.cell(row=fr, column=col, value=v)
            c.font = F_IN
            c.fill = FILL_IN
            c.number_format = GBP2 if col in (7, 10) else NUM1
        ws.cell(row=fr, column=6, value=0).number_format = GBP2
        ws.cell(row=fr, column=9, value=0).number_format = GBP2
        ws.cell(row=fr, column=12, value=0).number_format = NUM1
        ws.cell(row=fr, column=13, value=0).number_format = "0.00"
    else:
        ws.cell(row=fr, column=3,
                value=f"=IFERROR(AVERAGEIFS({prc}!$F${PF}:$F${PL},{prc}!$E${PF}:$E${PL},B{fr},{prc}!$AC${PF}:$AC${PL},1),0)"
                ).number_format = NUM1
        c = ws.cell(row=fr, column=4, value=None); c.font = F_IN; c.fill = FILL_IN; c.number_format = NUM1
        ws.cell(row=fr, column=6,
                value=f"=IFERROR(AVERAGEIFS({prc}!$I${PF}:$I${PL},{prc}!$E${PF}:$E${PL},B{fr},{prc}!$AC${PF}:$AC${PL},1),0)"
                ).number_format = GBP2
        c = ws.cell(row=fr, column=7, value=None); c.font = F_IN; c.fill = FILL_IN; c.number_format = GBP2
        # direct cost for P&L = consumables + machine energy + machine service (cols K,L,M)
        ws.cell(row=fr, column=9,
                value=(f"=IFERROR(AVERAGEIFS({prc}!$K${PF}:$K${PL},{prc}!$E${PF}:$E${PL},B{fr},{prc}!$AC${PF}:$AC${PL},1)"
                       f"+AVERAGEIFS({prc}!$L${PF}:$L${PL},{prc}!$E${PF}:$E${PL},B{fr},{prc}!$AC${PF}:$AC${PL},1)"
                       f"+AVERAGEIFS({prc}!$M${PF}:$M${PL},{prc}!$E${PF}:$E${PL},B{fr},{prc}!$AC${PF}:$AC${PL},1),0)")
                ).number_format = GBP2
        c = ws.cell(row=fr, column=10, value=None); c.font = F_IN; c.fill = FILL_IN; c.number_format = GBP2
        ws.cell(row=fr, column=12,
                value=f"=IFERROR(AVERAGEIFS({prc}!$G${PF}:$G${PL},{prc}!$E${PF}:$E${PL},B{fr},{prc}!$AC${PF}:$AC${PL},1),0)"
                ).number_format = NUM1
        mr_ = MACH["rows"].get(key)
        ws.cell(row=fr, column=13, value=(f"={mach_sheet}!$H${mr_}" if mr_ else 0)).number_format = "0.00"
    ws.cell(row=fr, column=5, value=f"=IF(D{fr}<>\"\",D{fr},C{fr})").number_format = NUM1
    ws.cell(row=fr, column=8, value=f"=IF(G{fr}<>\"\",G{fr},F{fr})").number_format = GBP2
    ws.cell(row=fr, column=11, value=f"=IF(J{fr}<>\"\",J{fr},I{fr})").number_format = GBP2
    mx = ws.cell(row=fr, column=14, value=seed_mix.get(key, 0) / 100)
    mx.font = F_IN
    mx.fill = FILL_IN
    mx.number_format = PCT1
    FAM_ROW[key] = fr
    fr += 1
f_last = fr - 1
ws.cell(row=fr, column=1, value="Mix check (must be 100%)").font = F_H2
mixcheck = ws.cell(row=fr, column=14, value=f"=SUM(N{f_first}:N{f_last})")
mixcheck.number_format = PCT1
mixcheck.font = F_WARN
fr += 1

# Monthly grid
g0 = fr + 2
def month_cols():
    return [(m, gcl(2 + m)) for m in range(1, MONTHS + 1)]  # data starts col C

ws.cell(row=g0, column=1, value="MONTHLY FORECAST").font = F_H2
ws.cell(row=g0, column=1).fill = FILL_SEC
g = g0 + 1
V = {}
def grid_row(gr, label, fmlfn, fmt, bold=False, fill=None, font=None):
    c = ws.cell(row=gr, column=1, value=label)
    c.font = F_H2 if bold else F_LBL
    for m, col in month_cols():
        cc = ws.cell(row=gr, column=2 + m, value=fmlfn(m, col))
        cc.font = font or F_FM
        cc.number_format = fmt
        if fill:
            cc.fill = fill
    return gr + 1

V["date"] = g
g = grid_row(g, "Month", lambda m, c: f"=EDATE({A['start']},COLUMN()-3)", "mmm yy", bold=True)
V["days"] = g
g = grid_row(g, "Days open", lambda m, c: f"=ROUND(DAY(EOMONTH({c}${V['date']},0))*{A['open_days']}/7,0)", NUM0)
V["roomhrs"] = g
g = grid_row(g, "Room-hours available", lambda m, c: f"={A['rooms']}*{A['open_hours']}*{c}${V['days']}", NUM0)
V["util"] = g
util_seed = [round(4 + (m - 1) * (25 - 4) / 35, 1) / 100 for m in range(1, MONTHS + 1)]
c = ws.cell(row=g, column=1, value="Utilisation % (EDIT — the model's biggest lever)")
c.font = F_H2
for m, col in month_cols():
    cc = ws.cell(row=g, column=2 + m, value=util_seed[m - 1])
    cc.font = F_IN
    cc.fill = FILL_IN
    cc.number_format = PCT1
g += 1
V["treated"] = g
g = grid_row(g, "Treated hours", lambda m, c: f"={c}${V['roomhrs']}*{c}${V['util']}", NUM0)
g += 1

# per family: hours, sessions, revenue, direct cost, machine kWh
fam_block = {}
for key in fam_order:
    frow = FAM_ROW[key]
    ws.cell(row=g, column=1, value=fam_labels[key]).font = F_H2
    ws.cell(row=g, column=1).fill = FILL_TOT
    g += 1
    rows = {}
    rows["hours"] = g
    g = grid_row(g, "  hours", lambda m, c, fr_=frow: f"={c}${V['treated']}*$N${fr_}", NUM1)
    rows["sessions"] = g
    g = grid_row(g, "  sessions", lambda m, c, fr_=frow, hr_=rows["hours"]: f"=IF($E${fr_}=0,0,{c}{hr_}*60/$E${fr_})", NUM1)
    rows["revenue"] = g
    g = grid_row(g, "  gross revenue", lambda m, c, fr_=frow, sr_=rows["sessions"]: f"={c}{sr_}*$H${fr_}", GBP0)
    rows["cost"] = g
    g = grid_row(g, "  direct costs", lambda m, c, fr_=frow, sr_=rows["sessions"]: f"={c}{sr_}*$K${fr_}", GBP0)
    rows["kwh"] = g
    g = grid_row(g, "  machine kWh", lambda m, c, fr_=frow, sr_=rows["sessions"]: f"={c}{sr_}*$L${fr_}/60*$M${fr_}", NUM1)
    fam_block[key] = rows

g += 1
V["tot_rev"] = g
rev_cells = lambda col: "+".join(f"{col}{fam_block[k]['revenue']}" for k in fam_order)
g = grid_row(g, "TOTAL gross booking revenue", lambda m, c: f"={rev_cells(c)}", GBP0, bold=True, fill=FILL_TOT)
V["tot_cost"] = g
cost_cells = lambda col: "+".join(f"{col}{fam_block[k]['cost']}" for k in fam_order)
g = grid_row(g, "TOTAL direct costs (consumables+machine energy+service)", lambda m, c: f"={cost_cells(c)}", GBP0, bold=True, fill=FILL_TOT)
V["tot_kwh"] = g
kwh_cells = lambda col: "+".join(f"{col}{fam_block[k]['kwh']}" for k in fam_order)
g = grid_row(g, "TOTAL machine kWh", lambda m, c: f"={kwh_cells(c)}", NUM0, bold=True, fill=FILL_TOT)
V["tot_sessions"] = g
ses_cells = lambda col: "+".join(f"{col}{fam_block[k]['sessions']}" for k in fam_order)
g = grid_row(g, "TOTAL sessions", lambda m, c: f"={ses_cells(c)}", NUM0, bold=True, fill=FILL_TOT)
V["treat_hrs_rev"] = g
hrs_rev = lambda col: "+".join(f"{col}{fam_block[k]['hours']}" for k in fam_order if k != "consultation")
g = grid_row(g, "Revenue-generating treated hours", lambda m, c: f"={hrs_rev(c)}", NUM0)

ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 4
for m in range(1, MONTHS + 1):
    ws.column_dimensions[gcl(2 + m)].width = 10
ws.freeze_panes = "C5"

# =====================================================================
# Energy
# =====================================================================
ws = wb.create_sheet("Energy")
ws["A1"] = "Energy — heating & cooling by season, base load, machine consumption"
ws["A1"].font = F_TITLE
ws["A2"] = energy["hvacBasis"]
ws["A2"].font = F_NOTE
ws["A3"] = "Sanity check from research: " + energy["annualSanityCheck"]
ws["A3"].font = F_NOTE

hv = 5
ws.cell(row=hv, column=1, value="HVAC kWh per room per month (EDIT — winter = heating via A/C heat pump, summer = cooling)").font = F_H2
hv += 1
months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
for i, mn in enumerate(months_names):
    ws.cell(row=hv, column=2 + i, value=mn).font = F_H2
hv += 1
for i, mn in enumerate(months_names):
    c = ws.cell(row=hv, column=2 + i, value=energy["hvacKWhPerRoomPerMonth"][mn])
    c.font = F_IN
    c.fill = FILL_IN
    c.number_format = NUM0
HVAC_CURVE = f"Energy!$B${hv}:$M${hv}"
ws.cell(row=hv, column=1, value="kWh/room").font = F_LBL

g0 = hv + 2
ws.cell(row=g0, column=1, value="MONTHLY ENERGY MODEL").font = F_H2
ws.cell(row=g0, column=1).fill = FILL_SEC
g = g0 + 1
E = {}
def erow(gr, label, fmlfn, fmt, bold=False, fill=None):
    ws.cell(row=gr, column=1, value=label).font = F_H2 if bold else F_LBL
    for m in range(1, MONTHS + 1):
        col = gcl(2 + m)
        cc = ws.cell(row=gr, column=2 + m, value=fmlfn(m, col))
        cc.font = F_FM
        cc.number_format = fmt
        if fill:
            cc.fill = fill
    return gr + 1

E["date"] = g
g = erow(g, "Month", lambda m, c: f"=EDATE({A['start']},COLUMN()-3)", "mmm yy", bold=True)
E["days"] = g
g = erow(g, "Days in month", lambda m, c: f"=DAY(EOMONTH({c}${E['date']},0))", NUM0)
E["hvac"] = g
g = erow(g, "HVAC kWh (rooms × seasonal curve)", lambda m, c: f"=INDEX({HVAC_CURVE},MONTH({c}${E['date']}))*{A['hvac_rooms']}", NUM0)
E["base"] = g
g = erow(g, "Base load kWh", lambda m, c: f"={A['baseload']}*{c}${E['days']}", NUM0)
E["machine"] = g
g = erow(g, "Machine kWh (from Volumes)", lambda m, c: f"=Volumes!{c}${V['tot_kwh']}", NUM0)
E["total_kwh"] = g
g = erow(g, "Total kWh", lambda m, c: f"={c}{E['hvac']}+{c}{E['base']}+{c}{E['machine']}", NUM0, bold=True, fill=FILL_TOT)
E["machine_cost"] = g
g = erow(g, "Machine energy cost £ (direct — already inside Pricing/Volumes direct costs)", lambda m, c: f"={c}{E['machine']}*{A['kwh_rate']}/100", GBP0)
E["facility_cost"] = g
g = erow(g, "Facility energy cost £ (HVAC + base + standing — to P&L overheads)",
         lambda m, c: f"=({c}{E['hvac']}+{c}{E['base']})*{A['kwh_rate']}/100+{c}{E['days']}*{A['standing']}/100", GBP0, bold=True, fill=FILL_TOT)
ws.column_dimensions["A"].width = 58
ws.column_dimensions["B"].width = 4
for m in range(1, MONTHS + 1):
    ws.column_dimensions[gcl(2 + m)].width = 9
ws.freeze_panes = "C6"

# =====================================================================
# P&L
# =====================================================================
ws = wb.create_sheet("P&L")
ws["A1"] = "P&L forecast — 36 months + annual summary"
ws["A1"].font = F_TITLE
ws["A2"] = ("Gross profit = net revenue − variable direct costs (consumables, machine energy & service, card fees). "
            "Salaried practitioners sit in overheads: they are capacity, not per-unit cost. "
            "Monthly corporation tax is a flat 19% accrual on positive months; the annual columns apply the real 19%/26.5%/25% bands. "
            "VAT switches on automatically when the rolling 12-month turnover tracker (bottom) crosses the threshold — prices are held, so VAT comes out of margin.")
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 44

P = {}
g = 4
YC = {1: gcl(3 + MONTHS), 2: gcl(4 + MONTHS), 3: gcl(5 + MONTHS)}  # annual cols AL/AM/AN

def prow(gr, label, fmlfn, fmt, bold=False, fill=None, yearfn="sum", indent=0):
    c = ws.cell(row=gr, column=1, value=("    " * indent) + label)
    c.font = F_H2 if bold else F_LBL
    for m in range(1, MONTHS + 1):
        col = gcl(2 + m)
        cc = ws.cell(row=gr, column=2 + m, value=fmlfn(m, col))
        cc.font = F_FM
        cc.number_format = fmt
        if fill:
            cc.fill = fill
    if yearfn == "sum":
        for y in (1, 2, 3):
            c0, c1 = gcl(2 + (y - 1) * 12 + 1), gcl(2 + y * 12)
            cc = ws.cell(row=gr, column=ws.max_column if False else (2 + MONTHS + y), value=f"=SUM({c0}{gr}:{c1}{gr})")
            cc.font = F_H2 if bold else F_FM
            cc.number_format = fmt
            cc.fill = FILL_KPI
    return gr + 1

# header dates + annual labels
P["date"] = g
g = prow(g, "Month", lambda m, c: f"=EDATE({A['start']},COLUMN()-3)", "mmm yy", bold=True, yearfn=None)
for y in (1, 2, 3):
    cc = ws.cell(row=P["date"], column=2 + MONTHS + y, value=f"Year {y}")
    cc.font = F_H1
    cc.fill = FILL_H1

P["gross_rev"] = g
g = prow(g, "Gross booking revenue", lambda m, c: f"=Volumes!{c}${V['tot_rev']}", GBP0, bold=True)
P["vat_flag_ref"] = None  # tracker at bottom; set later — use forward reference rows
# reserve tracker rows now (we know their positions at bottom); simpler: compute flag rows first at bottom AFTER, but formulas need refs — precompute row numbers:
# We'll lay out: main P&L block rows g.., then tracker block. To reference tracker rows in VAT row, compute expected offsets after building list count.
# Instead: put VAT tracker immediately here as hidden-ish rows? Cleanest: tracker before VAT row.
P["cum_rev"] = g
g = prow(g, "  cumulative gross revenue", lambda m, c: (f"={c}{P['gross_rev']}" if m == 1 else f"={gcl(1+m)}{P['cum_rev']}+{c}{P['gross_rev']}"), GBP0, yearfn=None, indent=1)
P["roll12"] = g
g = prow(g, "  rolling 12-mo taxable turnover (incl. pre-model estimate)",
         lambda m, c: (f"={c}{P['cum_rev']}+{A['vat_prior']}*(12-(COLUMN()-2))/12" if m < 12 else
                       (f"={c}{P['cum_rev']}" if m == 12 else f"={c}{P['cum_rev']}-{gcl(2+m-12)}{P['cum_rev']}")),
         GBP0, yearfn=None, indent=1)
P["vat_on"] = g
g = prow(g, "  VAT-registered this month? (1/0)",
         lambda m, c: (f"=IF({A['vat_override']}>0,IF(COLUMN()-2>={A['vat_override']},1,0),0)" if m == 1 else
                       f"=IF({A['vat_override']}>0,IF(COLUMN()-2>={A['vat_override']},1,0),IF(OR({gcl(1+m)}{P['vat_on']}=1,{gcl(1+m)}{P['roll12']}>={A['vat_thresh']}),1,0))"),
         NUM0, yearfn=None, indent=1)
P["vat_out"] = g
g = prow(g, "Output VAT (prices held — VAT absorbed)",
         lambda m, c: f"=-{c}{P['vat_flag'] if False else P['vat_on']}*{c}{P['gross_rev']}*{A['vat_rate']}/(1+{A['vat_rate']})", GBP0)
P["net_rev"] = g
g = prow(g, "Net revenue", lambda m, c: f"={c}{P['gross_rev']}+{c}{P['vat_out']}", GBP0, bold=True, fill=FILL_TOT)

P["cogs_direct"] = g
g = prow(g, "Consumables, machine energy & service", lambda m, c: f"=-Volumes!{c}${V['tot_cost']}", GBP0, indent=1)
P["cogs_card"] = g
g = prow(g, "Card processing fees", lambda m, c: f"=-{c}{P['gross_rev']}*{A['card_fee']}*{A['card_share']}", GBP0, indent=1)
P["gp"] = g
g = prow(g, "GROSS PROFIT", lambda m, c: f"={c}{P['net_rev']}+{c}{P['cogs_direct']}+{c}{P['cogs_card']}", GBP0, bold=True, fill=FILL_TOT)
P["gp_pct"] = g
g = prow(g, "Gross margin %", lambda m, c: f"=IF({c}{P['net_rev']}=0,0,{c}{P['gp']}/{c}{P['net_rev']})", PCT1, yearfn=None)
for y in (1, 2, 3):
    c0, c1 = gcl(2 + (y - 1) * 12 + 1), gcl(2 + y * 12)
    yc = gcl(2 + MONTHS + y)
    cc = ws.cell(row=P["gp_pct"], column=2 + MONTHS + y,
                 value=f"=IF(SUM({c0}{P['net_rev']}:{c1}{P['net_rev']})=0,0,SUM({c0}{P['gp']}:{c1}{P['gp']})/SUM({c0}{P['net_rev']}:{c1}{P['net_rev']}))")
    cc.number_format = PCT1
    cc.fill = FILL_KPI

# overheads
yearidx = lambda m: (m - 1) // 12 + 1
P["staff"] = g
def staff_f(m, c):
    y = yearidx(m)
    prac = {1: A['hc_prac_y1'], 2: A['hc_prac_y2'], 3: A['hc_prac_y3']}[y]
    rec = {1: A['hc_rec_y1'], 2: A['hc_rec_y2'], 3: A['hc_rec_y3']}[y]
    return (f"=-({prac}*{A['sal_prac']}+{rec}*{A['sal_rec']}+{A['sal_owner']})*{A['oncost']}/12")
g = prow(g, "Staff (practitioners, reception, owner)", staff_f, GBP0, indent=1)
P["rent"] = g
g = prow(g, "Rent", lambda m, c: f"=-{A['rent']}/12", GBP0, indent=1)
P["rates"] = g
g = prow(g, "Business rates", lambda m, c: f"=-{A['rates']}/12", GBP0, indent=1)
P["svc"] = g
g = prow(g, "Service charge", lambda m, c: f"=-{A['svc_charge']}/12", GBP0, indent=1)
P["energy"] = g
g = prow(g, "Facility energy (HVAC, base load, standing)", lambda m, c: f"=-Energy!{c}${E['facility_cost']}", GBP0, indent=1)
P["marketing"] = g
def mkt_f(m, c):
    y = yearidx(m)
    pct = {1: A['mkt_y1'], 2: A['mkt_y2'], 3: A['mkt_y3']}[y]
    return f"=-MAX({A['mkt_floor']},{pct}*{c}{P['net_rev']})"
g = prow(g, "Marketing", mkt_f, GBP0, indent=1)
P["insurance"] = g
g = prow(g, "Insurance", lambda m, c: f"=-{A['insurance']}/12", GBP0, indent=1)
P["software"] = g
g = prow(g, "Platform hosting & infra", lambda m, c: f"=-{A['software']}", GBP0, indent=1)
P["cleaning"] = g
g = prow(g, "Cleaning & sundries", lambda m, c: f"=-{A['cleaning']}", GBP0, indent=1)
P["other"] = g
g = prow(g, "Other fixed", lambda m, c: f"=-{A['other_fixed']}", GBP0, indent=1)
P["ebitda"] = g
oh = [P["staff"], P["rent"], P["rates"], P["svc"], P["energy"], P["marketing"], P["insurance"], P["software"], P["cleaning"], P["other"]]
g = prow(g, "EBITDA", lambda m, c: f"={c}{P['gp']}+" + "+".join(f"{c}{rr}" for rr in oh), GBP0, bold=True, fill=FILL_TOT)
P["ebitda_pct"] = g
g = prow(g, "EBITDA margin %", lambda m, c: f"=IF({c}{P['net_rev']}=0,0,{c}{P['ebitda']}/{c}{P['net_rev']})", PCT1, yearfn=None)
for y in (1, 2, 3):
    c0, c1 = gcl(2 + (y - 1) * 12 + 1), gcl(2 + y * 12)
    cc = ws.cell(row=P["ebitda_pct"], column=2 + MONTHS + y,
                 value=f"=IF(SUM({c0}{P['net_rev']}:{c1}{P['net_rev']})=0,0,SUM({c0}{P['ebitda']}:{c1}{P['ebitda']})/SUM({c0}{P['net_rev']}:{c1}{P['net_rev']}))")
    cc.number_format = PCT1
    cc.fill = FILL_KPI

P["dep"] = g
g = prow(g, "Depreciation (machinery)", lambda m, c: f"=-Machinery!{gcl(1+m)}${MACH['dep_row']}", GBP0, indent=1)
P["amort"] = g
g = prow(g, "Amortisation (enhancements/fit-out)", lambda m, c: f"=-Machinery!{gcl(1+m)}${MACH['amort_row']}", GBP0, indent=1)
P["ebit"] = g
g = prow(g, "EBIT", lambda m, c: f"={c}{P['ebitda']}+{c}{P['dep']}+{c}{P['amort']}", GBP0, bold=True)
P["interest"] = g
g = prow(g, "Interest / finance costs (EDIT if machines financed)", lambda m, c: f"=0", GBP0, indent=1)
for m in range(1, MONTHS + 1):
    cc = ws.cell(row=P["interest"], column=2 + m, value=0)
    cc.font = F_IN
    cc.fill = FILL_IN
    cc.number_format = GBP0
P["pbt"] = g
g = prow(g, "Profit before tax", lambda m, c: f"={c}{P['ebit']}+{c}{P['interest']}", GBP0, bold=True)
P["tax"] = g
g = prow(g, "Corporation tax (monthly ≈ 19% accrual; annual = banded)",
         lambda m, c: f"=-MAX(0,{c}{P['pbt']})*{A['ct_small']}", GBP0, yearfn=None, indent=1)
for y in (1, 2, 3):
    c0, c1 = gcl(2 + (y - 1) * 12 + 1), gcl(2 + y * 12)
    p_ = f"SUM({c0}{P['pbt']}:{c1}{P['pbt']})"
    cc = ws.cell(row=P["tax"], column=2 + MONTHS + y,
                 value=(f"=-IF({p_}<=0,0,IF({p_}<={A['ct_small_lim']},{p_}*{A['ct_small']},"
                        f"IF({p_}<={A['ct_marg_lim']},{A['ct_small_lim']}*{A['ct_small']}+({p_}-{A['ct_small_lim']})*{A['ct_marg']},{p_}*{A['ct_main']})))"))
    cc.number_format = GBP0
    cc.fill = FILL_KPI
P["np"] = g
g = prow(g, "NET PROFIT", lambda m, c: f"={c}{P['pbt']}+{c}{P['tax']}", GBP0, bold=True, fill=FILL_KPI, yearfn=None)
for y in (1, 2, 3):
    c0, c1 = gcl(2 + (y - 1) * 12 + 1), gcl(2 + y * 12)
    ycol = gcl(2 + MONTHS + y)
    cc = ws.cell(row=P["np"], column=2 + MONTHS + y, value=f"={ycol}{P['pbt']}+{ycol}{P['tax']}")
    cc.number_format = GBP0
    cc.fill = FILL_KPI
    cc.font = F_H2
P["np_pct"] = g
g = prow(g, "Net margin %", lambda m, c: f"=IF({c}{P['net_rev']}=0,0,{c}{P['np']}/{c}{P['net_rev']})", PCT1, yearfn=None)

g += 1
P["cap_need"] = g
g = prow(g, "Practitioner-hours needed (revenue treatments)", lambda m, c: f"=Volumes!{c}${V['treat_hrs_rev']}", NUM0, yearfn=None)
P["cap_have"] = g
def cap_f(m, c):
    y = yearidx(m)
    prac = {1: A['hc_prac_y1'], 2: A['hc_prac_y2'], 3: A['hc_prac_y3']}[y]
    return f"=({prac}+1)*{A['prac_cap']}"
g = prow(g, "Practitioner-hours available (headcount + owner)", cap_f, NUM0, yearfn=None)
P["cap_flag"] = g
g = prow(g, "CAPACITY CHECK", lambda m, c: f"=IF({c}{P['cap_need']}>{c}{P['cap_have']},\"OVER\",\"ok\")", "General", yearfn=None)
for m in range(1, MONTHS + 1):
    ws.cell(row=P["cap_flag"], column=2 + m).font = F_WARN

ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 4
for m in range(1, MONTHS + 1):
    ws.column_dimensions[gcl(2 + m)].width = 10
for y in (1, 2, 3):
    ws.column_dimensions[gcl(2 + MONTHS + y)].width = 12
ws.freeze_panes = "C5"

# =====================================================================
# Data
# =====================================================================
ws = wb.create_sheet("Data")
ws["A1"] = "Raw catalogue extract — audit trail for the Pricing tab"
ws["A1"].font = F_TITLE
ws["A2"] = f"Source: live production database via authenticated admin scrape, {EXTRACT_DATE}. Prices in pence as stored. costPence is NULL on every variant (no recorded COGS)."
ws["A2"].font = F_NOTE
hdr = ["Service slug", "Service", "Category", "vatClass (stored)", "Service status", "Variant", "durationMin",
       "displayDurationMin", "pricePence", "costPence", "courses (JSON)", "Variant status"]
for i, h in enumerate(hdr, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = F_H1
    c.fill = FILL_H1
dr = 5
for s in cat["services"]:
    for v in s.get("variants", []):
        vals = [s.get("slug"), s.get("name"), s.get("category"), s.get("vatClass"), s.get("status"),
                v.get("name"), v.get("durationMin"), v.get("displayDurationMin"), v.get("pricePence"),
                v.get("costPence"), json.dumps(v.get("courses")) if v.get("courses") else "", v.get("status") or ""]
        for i, val in enumerate(vals, 1):
            ws.cell(row=dr, column=i, value=val).font = F_NOTE
        dr += 1
dr += 2
ws.cell(row=dr, column=1, value="Booking mix — last 90 days (post-reopening)").font = F_H2
dr += 1
ws.cell(row=dr, column=1, value="treatmentSlug").font = F_H2
ws.cell(row=dr, column=2, value="bookings90d").font = F_H2
dr += 1
for b in cat["bookingStats"]["byTreatment"]:
    ws.cell(row=dr, column=1, value=b["treatmentSlug"]).font = F_NOTE
    ws.cell(row=dr, column=2, value=b["bookings90d"]).font = F_NOTE
    dr += 1
for col, w in zip("ABCDEFGHIJKL", (24, 34, 11, 13, 14, 38, 11, 14, 10, 9, 40, 12)):
    ws.column_dimensions[col].width = w

# =====================================================================
# README
# =====================================================================
ws = wb.create_sheet("README", 0)
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 120
lines = [
    ("KClinics Financial Model", "title"),
    (f"Built {EXTRACT_DATE} from the live K-Clinics catalogue + UK 2026 cost research. This workbook is the basis for refactoring pricing and cost structure for the brand — it does not touch the admin.", "note"),
    ("", None),
    ("HOW TO USE", "h"),
    ("Blue text on a cream background = an input you edit. Everything else is a formula — change an input and the whole model (prices, energy, P&L) recalculates.", "b"),
    ("The three levers that matter most: the utilisation ramp (Volumes, blue row), the target margins by machine class (Pricing, top), and the machinery purchase prices (Machinery — replace class-typical estimates with your invoices).", "b"),
    ("", None),
    ("TABS", "h"),
    ("Assumptions — every global input with its source. • Machinery — machines, enhancements, depreciation & amortisation schedule. • Energy — seasonal A/C vs heating curve per room, base load, machine kWh, cost split. • Pricing — per-variant unit economics and recommended prices (the price architecture). • Volumes — capacity × utilisation × mix forecast. • P&L — 36 monthly columns + annual totals: gross profit, EBITDA, net profit. • Data — raw catalogue extract (audit trail).", "b"),
    ("", None),
    ("CONVENTIONS", "h"),
    ("Gross profit = net revenue − variable direct costs (consumables, machine energy & service, card fees). Salaried staff are overheads. EBITDA excludes depreciation & amortisation, which are charged below it. Monthly corporation tax is a flat 19% accrual; the annual columns apply the real 2026/27 bands (19% / 26.5% marginal / 25%). AIA means cash tax in a capex year will be lower than the accrual — the accounting P&L still depreciates over asset life.", "b"),
    ("VAT: the company is not VAT-registered. The P&L tracks rolling 12-month taxable turnover and switches VAT on the month after it crosses £90k (override on Assumptions). Prices are held constant at registration, so VAT comes out of margin — that is why the Pricing tab prices for the registered state by default.", "b"),
    ("", None),
    ("CATALOGUE ISSUES FOUND DURING EXTRACTION (fix in admin before publishing new prices)", "h"),
    ("• Laser Skin Resurfacing: the 5-session course (£450) is cheaper than the 3-session course (£480).", "b"),
    ("• Tattoo removal 'Small' shares course totals with 'Very Small'.", "b"),
    ("• 'Laser Hair Removal — Women' is flagged VAT-EXEMPT in the DB; the men's equivalent is standard-rated. Treated as a data error (modelled at 20%) — needs accountant sign-off.", "b"),
    ("• botox-r9w5 duplicates botox-j2j7 (excluded); Red Carpet Lift BB Glow duplicates the BB Glow service; three inactive services still carry prices.", "b"),
    ("", None),
    ("KEY GAPS — the model flags these as estimates; replace with real figures", "h"),
]
for gp in gaps:
    lines.append(("• " + gp, "b"))
lines += [
    ("", None),
    ("PROVENANCE", "h"),
    ("Catalogue & bookings: live production DB (read-only) via authenticated admin scrape, " + EXTRACT_DATE + ". Cost benchmarks: web research Aug 2026 — Ofgem/business tariff comparisons (energy), UK equipment vendors (machinery), UK clinic industry norms (staff, rent, marketing, insurance). Basis noted next to every assumption.", "b"),
]
rr = 2
for text, kind in lines:
    if kind is None:
        rr += 1
        continue
    c = ws.cell(row=rr, column=2, value=text)
    if kind == "title":
        c.font = F_TITLE
    elif kind == "h":
        c.font = F_H2
        c.fill = FILL_SEC
    elif kind == "note":
        c.font = F_NOTE
    else:
        c.font = F_LBL
        c.alignment = Alignment(wrap_text=True, vertical="top")
    rr += 1

wb.save(OUT)
print("saved", OUT)
print("pricing rows:", PRC)
print("volumes fam rows:", f_first, f_last, "grid start", g0)
print("P&L rows:", {k: v for k, v in P.items() if isinstance(v, int)})
