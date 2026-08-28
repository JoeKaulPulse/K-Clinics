#!/usr/bin/env python3
"""Produce an exact build spec for the Google Sheet from the verified workbook.

Output per sheet (spec_<sheet>.txt):
  FILL <range>  <anchor formula>     - anchor at range top-left; verified with
                                       openpyxl Translator that plain fill
                                       reproduces every stored formula exactly.
  SET <addr> <formula>               - formulas that are not fill-reproducible.
  BLOCK <range> (TSV lines follow)   - rectangular static values.
  VAL <addr> <value>                 - scattered static values.
"""
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter as gcl, column_index_from_string
import datetime, json, re

import os
HERE = os.path.dirname(os.path.abspath(__file__))
wb = load_workbook(os.path.join(HERE, 'KClinics-Financial-Model.xlsx'))

def fmt_val(v):
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%d')
    return v

for ws in wb.worksheets:
    formulas = {}
    statics = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if isinstance(c.value, str) and c.value.startswith('='):
                formulas[(c.row, c.column)] = c.value
            else:
                statics[(c.row, c.column)] = fmt_val(c.value)

    used = set()
    fills = []
    sets = []

    # vertical fills per column
    for (r, col) in sorted(formulas):
        if (r, col) in used:
            continue
        anchor = formulas[(r, col)]
        rr = r
        while (rr + 1, col) in formulas and (rr + 1, col) not in used:
            expected = Translator(anchor, origin=f"{gcl(col)}{r}").translate_formula(f"{gcl(col)}{rr+1}")
            if formulas[(rr + 1, col)] == expected:
                rr += 1
            else:
                break
        if rr > r:
            for x in range(r, rr + 1):
                used.add((x, col))
            fills.append((f"{gcl(col)}{r}:{gcl(col)}{rr}", anchor))
    # horizontal fills per row
    for (r, col) in sorted(formulas, key=lambda t: (t[0], t[1])):
        if (r, col) in used:
            continue
        anchor = formulas[(r, col)]
        cc = col
        while (r, cc + 1) in formulas and (r, cc + 1) not in used:
            expected = Translator(anchor, origin=f"{gcl(col)}{r}").translate_formula(f"{gcl(cc+1)}{r}")
            if formulas[(r, cc + 1)] == expected:
                cc += 1
            else:
                break
        if cc > col:
            for x in range(col, cc + 1):
                used.add((r, x))
            fills.append((f"{gcl(col)}{r}:{gcl(cc)}{r}", anchor))
    for (r, col) in sorted(formulas):
        if (r, col) not in used:
            sets.append((f"{gcl(col)}{r}", formulas[(r, col)]))

    # merge vertical fills of adjacent columns with same relative pattern into rectangles
    merged = []
    fills_by_span = {}
    for rng, anchor in fills:
        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", rng)
        c1, r1, c2, r2 = m.groups()
        fills_by_span.setdefault((int(r1), int(r2), c1 == c2), []).append((column_index_from_string(c1), column_index_from_string(c2), anchor, rng))
    for (r1, r2, is_vert), lst in fills_by_span.items():
        if not is_vert or r1 == r2:
            for c1, c2, anchor, rng in lst:
                merged.append((rng, anchor))
            continue
        lst.sort()
        i = 0
        while i < len(lst):
            c1, c2, anchor, rng = lst[i]
            endc = c1
            a = anchor
            j = i + 1
            while j < len(lst) and lst[j][0] == endc + 1:
                expected = Translator(a, origin=f"{gcl(c1)}{r1}").translate_formula(f"{gcl(lst[j][0])}{r1}")
                if lst[j][2] == expected:
                    endc = lst[j][0]
                    j += 1
                else:
                    break
            merged.append((f"{gcl(c1)}{r1}:{gcl(endc)}{r2}", anchor))
            i = j

    # static rectangular blocks: group contiguous cells per row, then merge equal row-spans
    row_runs = {}
    for (r, col) in sorted(statics):
        runs = row_runs.setdefault(r, [])
        if runs and runs[-1][1] == col - 1:
            runs[-1][1] = col
        else:
            runs.append([col, col])
    blocks = []
    consumed = set()
    for r in sorted(row_runs):
        for c1, c2 in row_runs[r]:
            if (r, c1) in consumed:
                continue
            rr = r
            while True:
                nxt = rr + 1
                if nxt in row_runs and [c1, c2] in [list(x) for x in row_runs[nxt]] and (nxt, c1) not in consumed:
                    rr = nxt
                else:
                    break
            for x in range(r, rr + 1):
                consumed.add((x, c1))
            blocks.append((r, rr, c1, c2))

    out = []
    for rng, anchor in sorted(merged, key=lambda t: (int(re.search(r"(\d+)", t[0]).group(1)),)):
        out.append(f"FILL\t{rng}\t{anchor}")
    for addr, f in sets:
        out.append(f"SET\t{addr}\t{f}")
    for r, rr, c1, c2 in blocks:
        if r == rr and c1 == c2:
            out.append(f"VAL\t{gcl(c1)}{r}\t{statics[(r, c1)]}")
        else:
            out.append(f"BLOCK\t{gcl(c1)}{r}:{gcl(c2)}{rr}")
            for x in range(r, rr + 1):
                out.append("\t".join("" if (x, cc) not in statics else str(statics[(x, cc)]) for cc in range(c1, c2 + 1)))
    fn = os.path.join(HERE, "sheet-build-spec", f"spec_{ws.title.replace('&','n')}.txt")
    open(fn, 'w').write("\n".join(out))
    nfill = len(merged); nset = len(sets); nblk = len(blocks)
    import os
    print(f"{ws.title:12s} fills={nfill:3d} sets={nset:3d} static_blocks={nblk:3d} bytes={os.path.getsize(fn)}")
